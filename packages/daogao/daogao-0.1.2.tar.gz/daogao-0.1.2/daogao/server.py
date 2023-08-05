# -*- coding: utf8 -*-
from __future__ import print_function
from tornado.web import RequestHandler
from tornado.ioloop import IOLoop
import tornado.web
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from openpyxl.cell import Cell
from email.mime.text import MIMEText
from email.utils import formataddr
import ConfigParser as configparser
import requests
import sys
import os
import argparse
from binascii import hexlify
from threading import Thread
try:
    import simplejson as json
except ImportError:
    import json


hilos = []

unique_id = lambda: hexlify(os.urandom(16))
downloads_url = 'http://daogao.local'

def send_email(to=None, subject=None, content=""):
    api_key = "key-86dcacf19db8cf3716276f66992814d2"
    url_template = "https://api.mailgun.net/v3/{}/messages"
    domain = "edesarrollos.info"
    url = url_template.format(domain)
    response = requests.post(url,
        auth=("api", api_key,),
        data={
            "from": "postmaster@edesarrollos.info",
            "to": to,
            "subject": subject,
            "html": content
            }
        )

def process_xls(data):
    header = data['header']
    title = header['title']
    origin = data['dataOrigin']
    book = Workbook()
    sheet = book.active
    doc_id = unique_id()

    if 'logoURL' in header:
        try:
            response = requests.get(header['logoURL'], stream=True)
            logo = Image(response.raw)
            logo = Image(logo.image.resize((100, 100)))
        except requests.ConnectionError as cerror:
            print(cerror, file=sys.stderr)
            pass
    else:
        logo = None


    hdr_bkg_color = header['backgroundColor']
    header_bkg = PatternFill(fill_type="solid",
                             start_color=hdr_bkg_color,
                             end_color=hdr_bkg_color)

    if origin == 'array':
        rows = data['rows']

        cell = Cell(sheet, value=title)
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center')

        sheet.append(['', '', '', cell])

        sheet.merge_cells('A1:C1')
        sheet.merge_cells('D1:G1')

        for row in rows:
            sheet.append(row)

    else:
        db = data['database']
        sql_query = data['sqlQuery']
        url_callback = data['urlCallback']
        title = data['title']
        columns = data['columns']

        conn = pg_connect(host=db['host'],
                          database=db['name'],
                          password=db['password'],
                          user=db['user'])

        cursor = conn.cursor()
        cursor.execute(sql_query)

        index = 0

        is_first = True
        for row in cursor:
            if is_first:
                sheet.merge_cells('A1:C1')
                sheet.merge_cells('D1:G1')

                sheet.append(['', '', '', cell])

                if logo:
                    sheet.add_image(logo, 'A1')

                headcells = []
                for col in columns:
                    cell = Cell(sheet, value=col['label'])
                    cell.fill = header_bkg
                    headcells.append(cell)

                sheet.append(headcells)

                is_first = False
                #sheet.row_dimensions[0].height = 300
                sheet.row_dimensions[1].height = 72
            
            sheet.append(row)

            index += 1

    book.save('./{}.xlsx'.format(doc_id))

    return doc_id

def process_report(data):
    
    
    output = data['output']
    outtype = output['type']

    doc_id = None
    
    if outtype == 'xls':
        doc_id = process_xls(data)
    
    else:
        raise NotImplementedError()

    if 'triggers' in data:
        for trigger in data['triggers']:
            trigger_type = trigger['type']
            if trigger_type == 'sendEmail':
                filename = '{}.xlsx'.format(doc_id)
                filepath = '{}/{}'.format(downloads_url, filename)
                
                body_template = trigger.get(
                    'bodyTemplate', 'Este es tu archivo: {DownloadURL}'
                    )
                content = body_template.format(DownloadURL=filepath)
                subject = trigger.get('subject', 'Tu excel está listo')
                send_email(to=trigger['emails'],
                           subject=subject,
                           content=content,
                           )

class GenerationHandler(RequestHandler):

    def post(self):
        config = self.settings['config']
        api_method = config.get('api', 'method')
        headers = self.request.headers

        if api_method == 'singlekey':
            expected_api_key = config.get('api', 'api_key')
            api_key = headers.get('Authorization', None)

            if api_key == expected_api_key:
                args = json.loads(self.request.body)
                thr = Thread(target=process_report, args=(args, ))
                hilos.append(thr)
                thr.start()

                print("hay {} hilos".format(len(hilos)))

                self.write('OK')

            else:
                self.set_status(403)
                self.write('ERROR')


urls = [
    (r'^/generate-xls', GenerationHandler),
]

def main_loop():
    config_file = None
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', action='store', type=str,
                        dest='config_file', required=True)
    args = parser.parse_args()
    cparser = configparser.RawConfigParser()
    cparser.readfp(open(args.config_file, 'r'))

    try:
        app = tornado.web.Application(urls, config=cparser)
        loop = IOLoop()
        app.listen(8081)
        loop.start()
    except KeyboardInterrupt:
        print("bye")
    

if __name__ == '__main__':
    main_loop()